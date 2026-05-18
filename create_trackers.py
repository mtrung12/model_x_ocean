"""Create F1 Score (Weighted) and Retrieve Accuracy tracker Excel files."""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import statistics

# ── colour palette ──────────────────────────────────────────────────────────
NAVY    = "1F3864"   # dark header bg
BLUE    = "2E75B6"   # section header bg
LTBLUE  = "D6E4F0"   # alt row / sub-header
GREEN   = "E2EFDA"   # MYP section
YELLOW  = "FFF2CC"   # Essays section
PURPLE  = "EAD1DC"   # retrieve accuracy method header
GREY    = "F2F2F2"   # definition row bg
WHITE   = "FFFFFF"

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def make_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb

# ============================================================
#  FILE 1 : F1 Score Tracker (Weighted F1)
# ============================================================

# Data assembled from evaluation_summary.csv files
# Format: (model, method, run_id, dataset, {trait: weighted_f1})
F1_DATA = [
    # ── MYP (50 records) ─────────────────────────────────────
    ("gpt-4o-mini", "rag_oneshot",             "20260504-141756", "MYP", {
        "OPE": 0.5544, "CON": 0.3114, "EXT": 0.4338, "AGR": 0.4429, "NEU": 0.5774}),
    ("gpt-4o-mini", "cot_rag_oneshot",         "20260428-015151", "MYP", {
        "OPE": 0.5277, "CON": 0.3681, "EXT": 0.6136, "AGR": 0.5586, "NEU": 0.4067}),
    ("gpt-4o-mini", "rag_def_oneshot",         "20260504-213630", "MYP", {
        "OPE": 0.6064, "CON": 0.4678, "EXT": 0.4757, "AGR": 0.5106, "NEU": 0.5715}),
    ("gpt-4o-mini", "reasoned_rag_def_oneshot","20260518-041417", "MYP", {
        "OPE": 0.5725, "CON": 0.5757, "EXT": 0.1551, "AGR": 0.5202, "NEU": 0.3943}),
    ("llama3-8b",   "reasoned_rag_def_oneshot","20260517-175336", "MYP", {
        "OPE": 0.5829, "CON": 0.4732, "EXT": 0.1996, "AGR": 0.4337, "NEU": 0.5425}),

    # ── Essays (247 records) ──────────────────────────────────
    ("gpt-4o-mini", "oneshot",                 "20260424-141553", "Essays", {
        "OPE": 0.5504, "CON": 0.4483, "EXT": 0.5011, "AGR": 0.5082, "NEU": 0.5842}),
    ("gpt-4o-mini", "rag_oneshot",             "20260503-182339", "Essays", {
        "OPE": 0.5496, "CON": 0.3866, "EXT": 0.5861, "AGR": 0.5254, "NEU": 0.5374}),
    ("gpt-4o-mini", "reasoned_rag_def_oneshot","20260518-001847", "Essays", {
        "OPE": 0.5741, "CON": 0.5172, "EXT": 0.5387, "AGR": 0.5671, "NEU": 0.3890}),
]

METHOD_FULL = {
    "oneshot":                  "One-Shot (no RAG)",
    "rag_oneshot":              "RAG One-Shot",
    "cot_rag_oneshot":          "CoT + RAG One-Shot",
    "rag_def_oneshot":          "RAG + Definition One-Shot",
    "reasoned_rag_def_oneshot": "Reasoned RAG + Definition One-Shot",
}

TRAITS = ["OPE", "CON", "EXT", "AGR", "NEU"]
TRAIT_FULL = {
    "OPE": "Openness", "CON": "Conscientiousness",
    "EXT": "Extraversion", "AGR": "Agreeableness", "NEU": "Neuroticism",
}

def build_f1_sheet(ws, dataset_label, dataset_color, rows):
    """Populate one sheet of the F1 tracker."""
    ws.sheet_view.showGridLines = False

    # ── title row ──────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"F1 Score Tracker (Weighted F1) — {dataset_label} Dataset"
    c.font  = Font(bold=True, size=14, color="FFFFFF")
    c.fill  = solid(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── header row ─────────────────────────────────────────────
    headers = ["Model", "Method", "Run ID"] + TRAITS + ["AVG Weighted F1"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = header_font(size=10)
        c.fill      = solid(BLUE)
        c.alignment = center(wrap=True)
        c.border    = thin_border()
    ws.row_dimensions[2].height = 30

    # ── data rows ──────────────────────────────────────────────
    for ri, (model, method, run_id, _, traits) in enumerate(rows, 3):
        fill_color = dataset_color if ri % 2 == 0 else WHITE
        values = [
            model,
            METHOD_FULL.get(method, method),
            run_id,
        ] + [traits.get(t) for t in TRAITS]
        avg = statistics.mean(v for v in traits.values() if v is not None)
        values.append(round(avg, 4))

        for ci, v in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill   = solid(fill_color)
            c.border = thin_border()
            c.alignment = center()
            if isinstance(v, float):
                c.number_format = "0.0000"

    # ── column widths ──────────────────────────────────────────
    widths = [14, 32, 18, 10, 10, 10, 10, 10, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── freeze header ──────────────────────────────────────────
    ws.freeze_panes = "A3"


def build_f1_workbook():
    wb = make_wb()

    myp_rows    = [r for r in F1_DATA if r[3] == "MYP"]
    essays_rows = [r for r in F1_DATA if r[3] == "Essays"]

    ws_myp = wb.create_sheet("MYP (50 records)")
    build_f1_sheet(ws_myp, "MYP (50 records)", GREEN, myp_rows)

    ws_ess = wb.create_sheet("Essays (247 records)")
    build_f1_sheet(ws_ess, "Essays (247 records)", YELLOW, essays_rows)

    out = r"f:\std\GR\code\model_x_ocean\F1_score_tracker_weighted.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

# ============================================================
#  FILE 2 : Retrieve Accuracy Tracker
# ============================================================

RAG_METHODS = [
    {
        "name": "facet_embed",
        "display": "Facet Embed",
        "definition": (
            "Embedding-based retrieval per personality trait facet. "
            "Each trait dimension is embedded independently using a dense encoder "
            "(e.g. sentence-transformer), then the top-k most semantically similar "
            "documents are retrieved per facet. Combines facet-level retrieval results "
            "into a unified candidate set."
        ),
    },
    {
        "name": "hybrid_facet",
        "display": "Hybrid Facet",
        "definition": (
            "Hybrid (sparse + dense) retrieval per trait facet. "
            "BM25 (sparse keyword-based) and dense embedding scores are linearly combined: "
            "score = gamma * BM25 + (1 - gamma) * embedding_score. "
            "The best gamma per trait is selected by maximising hit-rate on the validation set. "
            "gamma=1.0 → pure BM25; gamma=0.0 → pure dense embedding."
        ),
        "best_gamma": {
            # (trait, k): gamma
            ("Openness to Experience", 3): 1.0,
            ("Openness to Experience", 5): 1.0,
            ("Conscientiousness",      3): 0.8,
            ("Conscientiousness",      5): 1.0,
            ("Extraversion",           3): 0.2,
            ("Extraversion",           5): 0.2,
            ("Agreeableness",          3): 0.2,
            ("Agreeableness",          5): 0.4,
            ("Neuroticism",            3): 1.0,
            ("Neuroticism",            5): 1.0,
        },
    },
    {
        "name": "profile",
        "display": "Profile Embed",
        "definition": (
            "Holistic user-profile embedding retrieval. "
            "A summary profile of the user (aggregated from all posts) is embedded as a single "
            "dense vector and used to retrieve the top-k most similar reference documents. "
            "No trait-specific decomposition; relies on the overall personality signal in the profile."
        ),
    },
    {
        "name": "rawpost",
        "display": "Raw Post Embed",
        "definition": (
            "Raw post embedding retrieval without personality-aware processing. "
            "Individual posts are embedded as-is and retrieved by cosine similarity. "
            "No trait faceting or profile aggregation—serves as a baseline that measures "
            "how well plain semantic similarity retrieves trait-relevant documents."
        ),
    },
    {
        "name": "sliced_dual",
        "display": "Sliced Dual",
        "definition": (
            "Dual-encoder retrieval with two strategies: "
            "(1) full_dual — retrieves using both the complete profile embedding and raw-post embedding "
            "simultaneously, merging result sets; "
            "(2) sliced_dual_asym — asymmetric variant where the query side uses sliced/windowed chunks "
            "of text rather than the full document, allowing finer-grained matching within long posts."
        ),
    },
]

# Retrieve accuracy data (from accuracy_summary CSVs, Essays = 247 records)
# Format: {method_name: {strategy: {(trait, k): {mean_match_rate, hit_rate}}}}

TRAIT_MAP_LONG = {
    "OPE": "Openness to Experience",
    "CON": "Conscientiousness",
    "EXT": "Extraversion",
    "AGR": "Agreeableness",
    "NEU": "Neuroticism",
}

# facet_embed
FACET_EMBED = {
    ("OPE", 3): (0.5358, 0.7045),
    ("CON", 3): (0.5830, 0.8178),
    ("EXT", 3): (0.5304, 0.8016),
    ("AGR", 3): (0.5587, 0.7733),
    ("NEU", 3): (0.5614, 0.7611),
    ("OPE", 5): (0.5377, 0.7692),
    ("CON", 5): (0.5814, 0.8866),
    ("EXT", 5): (0.5393, 0.8866),
    ("AGR", 5): (0.5482, 0.8462),
    ("NEU", 5): (0.5676, 0.8381),
}

# hybrid_facet – best gamma rows only
HYBRID_FACET = {
    ("OPE", 3): (0.5803, 0.6437),  # gamma=1.0
    ("CON", 3): (0.5722, 0.7692),  # gamma=0.8
    ("EXT", 3): (0.5709, 0.7854),  # gamma=0.2
    ("AGR", 3): (0.5722, 0.7773),  # gamma=0.2
    ("NEU", 3): (0.5965, 0.8543),  # gamma=1.0
    ("OPE", 5): (0.5676, 0.6964),  # gamma=1.0
    ("CON", 5): (0.5733, 0.8907),  # gamma=1.0
    ("EXT", 5): (0.5652, 0.8583),  # gamma=0.2
    ("AGR", 5): (0.5660, 0.8259),  # gamma=0.4
    ("NEU", 5): (0.5854, 0.8988),  # gamma=1.0
}

HYBRID_GAMMA = {
    ("OPE", 3): 1.0, ("OPE", 5): 1.0,
    ("CON", 3): 0.8, ("CON", 5): 1.0,
    ("EXT", 3): 0.2, ("EXT", 5): 0.2,
    ("AGR", 3): 0.2, ("AGR", 5): 0.4,
    ("NEU", 3): 1.0, ("NEU", 5): 1.0,
}

PROFILE = {
    ("OPE", 3): (0.5236, 0.8583),
    ("CON", 3): (0.5209, 0.8704),
    ("EXT", 3): (0.5209, 0.8583),
    ("AGR", 3): (0.5466, 0.8259),
    ("NEU", 3): (0.5574, 0.8300),
    ("OPE", 5): (0.5368, 0.9312),
    ("CON", 5): (0.5012, 0.9231),
    ("EXT", 5): (0.5336, 0.9352),
    ("AGR", 5): (0.5538, 0.9433),
    ("NEU", 5): (0.5709, 0.9312),
}

RAWPOST = {
    ("OPE", 3): (0.5020, 0.8826),
    ("CON", 3): (0.5236, 0.9069),
    ("EXT", 3): (0.5452, 0.8664),
    ("AGR", 3): (0.5169, 0.8907),
    ("NEU", 3): (0.5439, 0.8988),
    ("OPE", 5): (0.5150, 0.9717),
    ("CON", 5): (0.5109, 0.9798),
    ("EXT", 5): (0.5312, 0.9798),
    ("AGR", 5): (0.5198, 0.9879),
    ("NEU", 5): (0.5320, 0.9555),
}

# sliced_dual: two strategies
SLICED_DUAL = {
    "full_dual": {
        ("OPE", 3): (0.5169, 0.8907),
        ("CON", 3): (0.5155, 0.8826),
        ("EXT", 3): (0.5209, 0.8988),
        ("AGR", 3): (0.5358, 0.9190),
        ("NEU", 3): (0.5047, 0.8462),
        ("OPE", 5): (0.5174, 0.9636),
        ("CON", 5): (0.5304, 0.9717),
        ("EXT", 5): (0.5142, 0.9636),
        ("AGR", 5): (0.5401, 0.9879),
        ("NEU", 5): (0.5117, 0.9433),
    },
    "sliced_dual_asym": {
        ("OPE", 3): (0.5061, 0.8462),
        ("CON", 3): (0.5115, 0.8340),
        ("EXT", 3): (0.5277, 0.8704),
        ("AGR", 3): (0.5290, 0.8947),
        ("NEU", 3): (0.5385, 0.8664),
        ("OPE", 5): (0.5198, 0.9595),
        ("CON", 5): (0.5255, 0.9595),
        ("EXT", 5): (0.5174, 0.9514),
        ("AGR", 5): (0.5385, 0.9960),
        ("NEU", 5): (0.5368, 0.9433),
    },
}


def build_retrieve_workbook():
    wb = make_wb()
    ws = wb.create_sheet("Retrieve Accuracy")
    ws.sheet_view.showGridLines = False

    # ── column layout ──────────────────────────────────────────────────────
    # A: RAG Method | B: Strategy/Variant | C: Definition
    # D: Trait | E: k | F: Mean Match Rate | G: Hit Rate | H: Note (gamma)
    COL_METHOD   = 1
    COL_STRATEGY = 2
    COL_DEF      = 3
    COL_TRAIT    = 4
    COL_K        = 5
    COL_MMR      = 6
    COL_HR       = 7
    COL_NOTE     = 8

    widths = [18, 18, 60, 22, 5, 18, 12, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── main title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Retrieve Accuracy Tracker — Essays Dataset (n = 247 queries)"
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.fill  = solid(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── column headers ─────────────────────────────────────────────────────
    col_headers = [
        "RAG Method", "Strategy / Variant", "Definition",
        "Trait", "k", "Mean Match Rate", "Hit Rate", "Note",
    ]
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = header_font(size=10)
        c.fill      = solid(BLUE)
        c.alignment = center(wrap=True)
        c.border    = thin_border()
    ws.row_dimensions[2].height = 32

    row = 3

    def write_row(ws, r, method_name, strategy, definition,
                  trait, k, mmr, hr, note="", method_fill=WHITE, is_def=False):
        vals = [method_name, strategy, definition, trait, k,
                mmr, hr, note]
        fills = [method_fill, method_fill, GREY if is_def else WHITE,
                 WHITE, WHITE, WHITE, WHITE, WHITE]
        for ci, (v, f) in enumerate(zip(vals, fills), 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill   = solid(f)
            c.border = thin_border()
            c.alignment = center(wrap=True) if ci == 3 else center()
            if isinstance(v, float):
                c.number_format = "0.0000"

    # ── write each RAG method ──────────────────────────────────────────────
    method_fills = [LTBLUE, GREEN, YELLOW, PURPLE, "FCE4D6"]
    TRAIT_ORDER = ["OPE", "CON", "EXT", "AGR", "NEU"]

    for mi, method in enumerate(RAG_METHODS):
        mfill = method_fills[mi % len(method_fills)]
        mname = method["display"]
        mdef  = method["definition"]

        if method["name"] == "sliced_dual":
            strategies = ["full_dual", "sliced_dual_asym"]
            data_map = SLICED_DUAL
        elif method["name"] == "facet_embed":
            strategies = [None]
            data_map = {None: FACET_EMBED}
        elif method["name"] == "hybrid_facet":
            strategies = [None]
            data_map = {None: HYBRID_FACET}
        elif method["name"] == "profile":
            strategies = [None]
            data_map = {None: PROFILE}
        else:  # rawpost
            strategies = [None]
            data_map = {None: RAWPOST}

        # definition row (merged)
        ws.merge_cells(
            start_row=row, start_column=COL_METHOD,
            end_row=row,   end_column=COL_NOTE
        )
        c = ws.cell(row=row, column=COL_METHOD,
                    value=f"▶  {mname}   —   {mdef}")
        c.font      = Font(bold=True, italic=True, size=9, color="333333")
        c.fill      = solid(GREY)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()
        ws.row_dimensions[row].height = 48
        row += 1

        for strat in strategies:
            data = data_map[strat] if isinstance(data_map, dict) and strat in data_map \
                   else data_map.get(strat, data_map)

            strat_label = strat if strat else ""

            for trait in TRAIT_ORDER:
                for k in [3, 5]:
                    key = (trait, k)
                    mmr, hr = data.get(key, (None, None))
                    note = ""
                    if method["name"] == "hybrid_facet":
                        g = HYBRID_GAMMA.get(key)
                        if g is not None:
                            note = f"γ={g}"

                    write_row(ws, row,
                              mname if (strat == strategies[0] and trait == TRAIT_ORDER[0] and k == 3) else "",
                              strat_label if (trait == TRAIT_ORDER[0] and k == 3) else "",
                              "",
                              TRAIT_FULL[trait], k, mmr, hr, note,
                              method_fill=mfill)
                    row += 1

        # blank separator
        ws.row_dimensions[row].height = 6
        row += 1

    ws.freeze_panes = "A3"

    out = r"f:\std\GR\code\model_x_ocean\Retrieve_accuracy_tracker.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    build_f1_workbook()
    build_retrieve_workbook()
    print("Done.")
